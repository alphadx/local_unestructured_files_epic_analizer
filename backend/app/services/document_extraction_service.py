from __future__ import annotations

from dataclasses import dataclass
import logging
import re
from pathlib import Path

from app.models.schemas import ArtifactKind, DocumentChunk, FileIndex

logger = logging.getLogger(__name__)

_TEXT_EXTENSIONS = {".txt", ".csv", ".md", ".json", ".xml", ".html", ".htm"}
_SPREADSHEET_EXTENSIONS = {".xlsx", ".xlsm"}
_PDF_EXTENSIONS = {".pdf"}
_DOCX_EXTENSIONS = {".docx"}
_CONTAINER_TEXTUAL_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".pptx",
    ".xlsx",
    ".xlsm",
}
_MAX_CHUNK_CHARS = 1_200
_MAX_TEXT_CHARS = 20_000
_MAX_CLASSIFICATION_CHARS = 12_000

# Binary MIME type prefixes that should be skipped without attempting extraction
_BINARY_MIME_PREFIXES = {
    "image/",
    "video/",
    "audio/",
    "application/x-executable",
    "application/x-sharedlib",
    "application/x-dvi",
    "application/x-jar",
    "application/gzip",
    "application/x-gzip",
    "application/zip",
    "application/x-rar-compressed",
    "application/x-7z-compressed",
    "application/x-tar",
    "application/octet-stream",
}

# Binary file extensions to skip
_BINARY_EXTENSIONS = {
    ".exe", ".dll", ".so", ".dylib", ".bin", ".app", ".msi",
    ".jar", ".com", ".bat", ".cmd", ".sh", ".class",
    ".o", ".a", ".lib", ".pyc", ".pyo",
    ".zip", ".tar", ".gz", ".gzip", ".rar", ".7z", ".bz2", ".xz",
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".ico",
    ".mp3", ".mp4", ".avi", ".mov", ".mkv", ".flv", ".wav",
}


@dataclass(slots=True)
class DocumentExtraction:
    documento_id: str
    source_path: str
    text: str
    chunks: list[DocumentChunk]
    extraction_method: str
    element_count: int = 0


def _read_text_file(path: Path) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        return fh.read(_MAX_TEXT_CHARS)


def _normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def _chunk_text(
    documento_id: str,
    source_path: str,
    text: str,
    *,
    title: str | None = None,
    section_path: list[str] | None = None,
    page_number: int | None = None,
) -> list[DocumentChunk]:
    cleaned = _normalize_text(text)
    if not cleaned:
        return []

    paragraphs = [p.strip() for p in cleaned.split("\n\n") if p.strip()]
    if not paragraphs:
        paragraphs = [cleaned]

    chunks: list[DocumentChunk] = []
    buffer = ""
    chunk_index = 0

    def _emit(current_text: str) -> None:
        nonlocal chunk_index
        normalized = _normalize_text(current_text)
        if not normalized:
            return
        chunks.append(
            DocumentChunk(
                chunk_id=f"{documento_id}::chunk::{chunk_index:04d}",
                documento_id=documento_id,
                artifact_kind=ArtifactKind.CHUNK,
                source_path=source_path,
                chunk_index=chunk_index,
                text=normalized,
                title=title,
                section_path=list(section_path or []),
                page_number=page_number,
                token_count=max(1, len(normalized.split())),
            )
        )
        chunk_index += 1

    for paragraph in paragraphs:
        if len(paragraph) <= _MAX_CHUNK_CHARS:
            candidate = f"{buffer}\n\n{paragraph}" if buffer else paragraph
            if len(candidate) <= _MAX_CHUNK_CHARS:
                buffer = candidate
                continue
            _emit(buffer)
            buffer = paragraph
            continue

        if buffer:
            _emit(buffer)
            buffer = ""

        start = 0
        while start < len(paragraph):
            _emit(paragraph[start : start + _MAX_CHUNK_CHARS])
            start += _MAX_CHUNK_CHARS

    if buffer:
        _emit(buffer)

    return chunks


def _extract_with_unstructured(path: Path, documento_id: str) -> tuple[str, list[DocumentChunk], int, str | None]:
    """
    Returns (text, chunks, count, failure_reason).
    failure_reason is None on success, or a short string explaining why extraction yielded nothing.
    """
    try:
        from unstructured.partition.auto import partition  # type: ignore
    except ImportError:
        return "", [], 0, "librería 'unstructured' no instalada"

    try:
        elements = partition(filename=str(path))
    except Exception as exc:  # noqa: BLE001
        logger.warning("unstructured partition failed for %s: %s", path, exc)
        return "", [], 0, f"{type(exc).__name__}: {exc}"

    raw_parts: list[str] = []
    chunks: list[DocumentChunk] = []

    for idx, element in enumerate(elements):
        text = getattr(element, "text", "") or ""
        cleaned = _normalize_text(text)
        if not cleaned:
            continue
        raw_parts.append(cleaned)

        metadata = getattr(element, "metadata", None)
        title = getattr(element, "category", None) or element.__class__.__name__
        page_number = getattr(metadata, "page_number", None) if metadata else None
        section_path = []
        if title:
            section_path.append(str(title))
        chunks.extend(
            _chunk_text(
                documento_id=documento_id,
                source_path=str(path),
                text=cleaned,
                title=str(title) if title else None,
                section_path=section_path,
                page_number=page_number,
            )
        )

    if not chunks:
        return "", [], 0, "sin elementos con texto"
    return "\n\n".join(raw_parts), chunks, len(raw_parts), None


def _extract_with_openpyxl(path: Path, documento_id: str) -> tuple[str, list[DocumentChunk], int, str | None]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return "", [], 0, "librería 'openpyxl' no instalada"

    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl failed for %s: %s", path, exc)
        return "", [], 0, f"{type(exc).__name__}: {exc}"

    raw_parts: list[str] = []

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if not values:
                    continue
                line = " | ".join(values)
                raw_parts.append(f"[{sheet.title}] {line}")
    finally:
        try:
            workbook.close()
        except Exception:  # noqa: BLE001
            pass

    workbook_text = "\n".join(raw_parts)
    chunks = _chunk_text(
        documento_id=documento_id,
        source_path=str(path),
        text=workbook_text,
        title=path.name,
        section_path=["xlsx"],
    )

    if not chunks:
        return "", [], 0, "sin celdas con texto"
    return workbook_text, chunks, len(raw_parts), None


def _extract_with_pypdf(path: Path, documento_id: str) -> tuple[str, list[DocumentChunk], int, str | None]:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        return "", [], 0, "librería 'pypdf' no instalada"

    try:
        reader = PdfReader(str(path))
    except Exception as exc:  # noqa: BLE001
        logger.warning("pypdf failed opening %s: %s", path, exc)
        return "", [], 0, f"{type(exc).__name__}: {exc}"

    raw_parts: list[str] = []
    chunks: list[DocumentChunk] = []

    for page_idx, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            logger.warning("pypdf failed extracting page %s (%s): %s", page_idx, path, exc)
            continue
        cleaned = _normalize_text(page_text)
        if not cleaned:
            continue
        raw_parts.append(cleaned)
        chunks.extend(
            _chunk_text(
                documento_id=documento_id,
                source_path=str(path),
                text=cleaned,
                title=path.name,
                section_path=["pdf"],
                page_number=page_idx + 1,
            )
        )

    if not chunks:
        return "", [], 0, "sin páginas con texto"
    return "\n\n".join(raw_parts), chunks, len(raw_parts), None


def _extract_with_docx(path: Path, documento_id: str) -> tuple[str, list[DocumentChunk], int, str | None]:
    try:
        import docx  # type: ignore
    except ImportError:
        return "", [], 0, "librería 'python-docx' no instalada"

    try:
        doc = docx.Document(str(path))
    except Exception as exc:  # noqa: BLE001
        logger.warning("python-docx failed opening %s: %s", path, exc)
        return "", [], 0, f"{type(exc).__name__}: {exc}"

    raw_parts: list[str] = []
    
    for para in doc.paragraphs:
        cleaned = _normalize_text(para.text)
        if not cleaned:
            continue
        raw_parts.append(cleaned)
        
    docx_text = "\n\n".join(raw_parts)
    chunks = _chunk_text(
        documento_id=documento_id,
        source_path=str(path),
        text=docx_text,
        title=path.name,
        section_path=["docx"],
    )

    if not chunks:
        return "", [], 0, "sin párrafos con texto"
    return docx_text, chunks, len(raw_parts), None


def _is_binary_file(file_index: FileIndex) -> bool:
    """Check if file is binary based on extension or MIME type."""
    # Check by extension first (fastest)
    ext_lower = file_index.extension.lower()
    if ext_lower in _BINARY_EXTENSIONS:
        return True
    
    # Check by MIME type prefix
    if file_index.mime_type:
        mime_lower = file_index.mime_type.lower()
        for binary_prefix in _BINARY_MIME_PREFIXES:
            if mime_lower.startswith(binary_prefix):
                # Some office/pdf files are misdetected as octet-stream by magic.
                if (
                    binary_prefix == "application/octet-stream"
                    and ext_lower in _CONTAINER_TEXTUAL_EXTENSIONS
                ):
                    continue
                return True
    
    return False


def extract_document_content(file_index: FileIndex) -> DocumentExtraction:
    """
    Extract readable content and semantic chunks from a file.

    Uses `unstructured` when available and falls back to basic text extraction
    for plain-text documents. Binary formats are detected early and skipped
    without attempting extraction.
    """
    path = Path(file_index.path)
    documento_id = file_index.sha256 or file_index.path

    text = ""
    chunks: list[DocumentChunk] = []
    extraction_method = "none"
    element_count = 0

    # Skip early if file is binary
    if _is_binary_file(file_index):
        logger.debug(
            "Skipping binary file (early detection): %s (ext=%s, mime=%s)",
            file_index.path,
            file_index.extension,
            file_index.mime_type or "unknown",
        )
        return DocumentExtraction(
            documento_id=documento_id,
            source_path=str(path),
            text="",
            chunks=[],
            extraction_method="skipped_binary",
            element_count=0,
        )

    unstructured_text, unstructured_chunks, count, unstructured_failure = _extract_with_unstructured(path, documento_id)
    if unstructured_chunks:
        text = unstructured_text
        chunks = unstructured_chunks
        extraction_method = "unstructured"
        element_count = count
    else:
        if unstructured_failure:
            extraction_method = f"none ({unstructured_failure})"
    if not chunks and file_index.extension.lower() in _SPREADSHEET_EXTENSIONS:
        xlsx_text, xlsx_chunks, xlsx_count, xlsx_failure = _extract_with_openpyxl(path, documento_id)
        if xlsx_chunks:
            text = xlsx_text
            chunks = xlsx_chunks
            extraction_method = "openpyxl"
            element_count = xlsx_count
        elif xlsx_failure:
            extraction_method = f"none ({xlsx_failure})"

    if not chunks and file_index.extension.lower() in _PDF_EXTENSIONS:
        pdf_text, pdf_chunks, pdf_count, pdf_failure = _extract_with_pypdf(path, documento_id)
        if pdf_chunks:
            text = pdf_text
            chunks = pdf_chunks
            extraction_method = "pypdf"
            element_count = pdf_count
        elif pdf_failure:
            extraction_method = f"none ({pdf_failure})"

    if not chunks and file_index.extension.lower() in _DOCX_EXTENSIONS:
        docx_text, docx_chunks, docx_count, docx_failure = _extract_with_docx(path, documento_id)
        if docx_chunks:
            text = docx_text
            chunks = docx_chunks
            extraction_method = "python-docx"
            element_count = docx_count
        elif docx_failure:
            extraction_method = f"none ({docx_failure})"

    if not chunks and (
        file_index.extension in _TEXT_EXTENSIONS
        or (file_index.mime_type and file_index.mime_type.startswith("text/"))
    ):
        try:
            text = _read_text_file(path)
            if text:
                chunks = _chunk_text(
                    documento_id=documento_id,
                    source_path=str(path),
                    text=text,
                    title=path.name,
                    section_path=[path.parent.as_posix()],
                )
                extraction_method = "text-file"
                element_count = len(chunks)
        except OSError as exc:
            logger.warning("Unable to read text file %s: %s", path, exc)
            text = ""

    if not chunks and text:
        chunks = _chunk_text(
            documento_id=documento_id,
            source_path=str(path),
            text=text,
            title=path.name,
            section_path=[path.parent.as_posix()],
        )
        element_count = len(chunks)
        if extraction_method == "none":
            extraction_method = "text-file"

    return DocumentExtraction(
        documento_id=documento_id,
        source_path=str(path),
        text=_normalize_text(text),
        chunks=chunks,
        extraction_method=extraction_method,
        element_count=element_count,
    )


def build_classification_context(extraction: DocumentExtraction) -> str:
    """
    Build a compact LLM input from extracted structure and text.

    The output keeps headings and chunk order while staying within a bounded
    character budget so the classifier sees the document shape, not just a raw
    blob of text.
    """
    lines: list[str] = []
    if extraction.extraction_method != "none":
        lines.append(f"Metodo de extraccion: {extraction.extraction_method}")

    if extraction.chunks:
        for chunk in extraction.chunks[:10]:
            label_parts = [f"chunk {chunk.chunk_index}"]
            if chunk.title:
                label_parts.append(chunk.title)
            if chunk.page_number is not None:
                label_parts.append(f"pagina {chunk.page_number}")
            heading = " | ".join(label_parts)
            body = _normalize_text(chunk.text)
            if len(body) > 2_000:
                body = body[:2_000].rstrip() + "..."
            lines.append(f"[{heading}]\n{body}")
    elif extraction.text:
        lines.append(extraction.text)

    context = "\n\n".join(lines).strip()
    return context[:_MAX_CLASSIFICATION_CHARS]
